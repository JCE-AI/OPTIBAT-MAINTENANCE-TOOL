import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# Definir los grupos encontrados
grupos_data = [
    {
        'Grupo': 'GRUPO 1',
        'Cantidad': 4,
        'Activos': 'ABG DHAR, ABG PALI, MOLINS ALION COLOMBIA, MOLINS-BCN-BARACELONA',
        'Total_Flags': 5,
        'Flags': 'OPTIBAT_ON, OPTIBAT_READY, OPTIBAT_SUPPORT, OPTIBAT_MACROSTATES, OPTIBAT_RESULTEXISTANCE',
        'Configuración': 'Estándar sin comunicación ni watchdog'
    },
    {
        'Grupo': 'GRUPO 2',
        'Cantidad': 3,
        'Activos': 'TITAN ALEXANDRIA CM7, CM8, CM9',
        'Total_Flags': 5,
        'Flags': 'OPTIBAT_ON, OPTIBAT_READY, OPTIBAT_COMMUNICATION, OPTIBAT_SUPPORT, OPTIBAT_MACROSTATES',
        'Configuración': 'Con comunicación, sin resultexistance'
    },
    {
        'Grupo': 'GRUPO 3',
        'Cantidad': 4,
        'Activos': 'TITAN-KOSJERIC (CM1, KILN, RM1), TITAN-PENNSUCO-VRM',
        'Total_Flags': 6,
        'Flags': 'OPTIBAT_ON, OPTIBAT_READY, OPTIBAT_COMMUNICATION, OPTIBAT_SUPPORT, OPTIBAT_MACROSTATES, OPTIBAT_RESULTEXISTANCE',
        'Configuración': 'Casi completo (falta watchdog)'
    },
    {
        'Grupo': 'GRUPO 4',
        'Cantidad': 2,
        'Activos': 'TITAN-PENNSUCO-FM3, TITAN-PENNSUCO-KILN',
        'Total_Flags': 6,
        'Flags': 'OPTIBAT_ON, OPTIBAT_READY, KILN_OPTIBAT_COMMUNICATION, OPTIBAT_SUPPORT, OPTIBAT_MACROSTATES, OPTIBAT_RESULTEXISTANCE',
        'Configuración': 'Comunicación especial KILN'
    },
    {
        'Grupo': 'GRUPO 5',
        'Cantidad': 3,
        'Activos': 'TITAN-ROANOKE (FM10, KILN, RM1)',
        'Total_Flags': 2,
        'Flags': 'OPTIBAT_READY, Communication_Flag',
        'Configuración': 'Mínima implementación'
    },
    {
        'Grupo': 'GRUPO 6',
        'Cantidad': 3,
        'Activos': 'TITAN-SHARR (CM2, KILN, RM2)',
        'Total_Flags': 1,
        'Flags': 'OPTIBAT_READY',
        'Configuración': 'Solo Ready flag'
    }
]

# Datos de configuraciones únicas
unicos_data = [
    {'Activo': 'ABG DALLA', 'Total_Flags': 7, 'Particularidad': 'IMPLEMENTACIÓN COMPLETA (todas las flags)'},
    {'Activo': 'CEMEX FM1 BALCONES', 'Total_Flags': 7, 'Particularidad': 'IMPLEMENTACIÓN COMPLETA (nomenclatura antigua _Copy)'},
    {'Activo': 'CRH LEMONA', 'Total_Flags': 5, 'Particularidad': 'Flag Ready especial: OPTIBAT_Ready_fromPLANT'},
    {'Activo': 'ANGUS', 'Total_Flags': 0, 'Particularidad': 'SIN IMPLEMENTACIÓN OPTIBAT'}
]

# Crear Excel
output_path = r'C:\Users\JuanCruz\Desktop_Local\mtto streamlit\STATISTICS FLAGS\RESUMEN_GRUPOS_ACTIVOS.xlsx'

with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
    # Hoja 1: Grupos con mismas flags
    df_grupos = pd.DataFrame(grupos_data)
    df_grupos.to_excel(writer, sheet_name='Grupos Idénticos', index=False)
    
    # Formatear hoja de grupos
    ws_grupos = writer.sheets['Grupos Idénticos']
    
    # Ajustar anchos
    ws_grupos.column_dimensions['A'].width = 12
    ws_grupos.column_dimensions['B'].width = 10
    ws_grupos.column_dimensions['C'].width = 60
    ws_grupos.column_dimensions['D'].width = 12
    ws_grupos.column_dimensions['E'].width = 80
    ws_grupos.column_dimensions['F'].width = 35
    
    # Estilos
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                   top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Aplicar formato a encabezados
    for cell in ws_grupos[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border
    
    # Colorear filas según cantidad de flags
    colors = {
        7: 'C6EFCE',  # Verde claro
        6: 'E2EFDA',  # Verde muy claro
        5: 'FFF2CC',  # Amarillo claro
        2: 'FFE6CC',  # Naranja claro
        1: 'F4CCCC'   # Rojo claro
    }
    
    for row in ws_grupos.iter_rows(min_row=2, max_row=ws_grupos.max_row):
        total_flags = row[3].value
        fill_color = colors.get(total_flags, 'FFFFFF')
        fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
        
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical='center', wrap_text=True)
            if cell.column in [2, 4]:  # Centrar columnas numéricas
                cell.alignment = center_align
            cell.fill = fill
    
    # Hoja 2: Configuraciones únicas
    df_unicos = pd.DataFrame(unicos_data)
    df_unicos.to_excel(writer, sheet_name='Configuraciones Únicas', index=False)
    
    ws_unicos = writer.sheets['Configuraciones Únicas']
    ws_unicos.column_dimensions['A'].width = 25
    ws_unicos.column_dimensions['B'].width = 15
    ws_unicos.column_dimensions['C'].width = 60
    
    # Formato encabezados
    for cell in ws_unicos[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border
    
    # Formato datos
    for row in ws_unicos.iter_rows(min_row=2, max_row=ws_unicos.max_row):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical='center', wrap_text=True)
            if cell.column == 2:
                cell.alignment = center_align
        
        # Colorear según flags
        total_flags = row[1].value
        if total_flags == 7:
            fill = PatternFill(start_color='00B050', end_color='00B050', fill_type='solid')
            row[1].font = Font(bold=True, color='FFFFFF')
        elif total_flags == 0:
            fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
            row[1].font = Font(bold=True, color='FFFFFF')
        else:
            fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
        
        for cell in row:
            cell.fill = fill
    
    # Hoja 3: Resumen estadístico
    resumen_stats = {
        'Métrica': [
            'Total de activos analizados',
            'Grupos con configuración idéntica',
            'Activos en grupos',
            'Activos con configuración única',
            'Grupo más grande',
            'Implementación más común'
        ],
        'Valor': [
            '23',
            '6 grupos',
            '19 activos (82.6%)',
            '4 activos (17.4%)',
            'Grupo 1 y 3 (4 activos cada uno)',
            '5 flags (7 activos)'
        ]
    }
    
    df_resumen = pd.DataFrame(resumen_stats)
    df_resumen.to_excel(writer, sheet_name='Resumen', index=False)
    
    ws_resumen = writer.sheets['Resumen']
    ws_resumen.column_dimensions['A'].width = 35
    ws_resumen.column_dimensions['B'].width = 40
    
    for cell in ws_resumen[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border
    
    for row in ws_resumen.iter_rows(min_row=2, max_row=ws_resumen.max_row):
        for cell in row:
            cell.border = border
            cell.alignment = center_align

print("Archivo Excel de grupos creado exitosamente")
print(f"Ubicación: {output_path}")
print("\nRESUMEN EJECUTIVO:")
print("- 19 de 23 activos (82.6%) comparten configuración con otros")
print("- 6 grupos identificados con configuraciones idénticas")
print("- Los grupos más grandes tienen 4 activos cada uno")
print("- Solo 4 activos tienen configuraciones únicas")