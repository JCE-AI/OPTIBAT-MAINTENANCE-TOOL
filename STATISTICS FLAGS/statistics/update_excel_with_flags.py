import pandas as pd
import os

# Dictionary with all the flag information
client_flags = {
    'ABG DALLA': {
        'OPTIBAT_ON': 'YES',
        'Flag_Ready': 'OPTIBAT_READY',
        'Communication_ECS': 'OPTIBAT_COMMUNICATION',
        'Support_Flag_Copy': 'OPTIBAT_SUPPORT',
        'Macrostates_Flag_Copy': 'OPTIBAT_MACROSTATES',
        'Resultexistance_Flag_Copy': 'OPTIBAT_RESULTS',
        'OPTIBAT_WATCHDOG': 'YES'
    },
    'ABG DHAR': {
        'OPTIBAT_ON': 'YES',
        'Flag_Ready': 'OPTIBAT_READY',
        'Communication_ECS': 'NO',
        'Support_Flag_Copy': 'OPTIBAT_SUPPORT',
        'Macrostates_Flag_Copy': 'OPTIBAT_MACROSTATES',
        'Resultexistance_Flag_Copy': 'OPTIBAT_RESULTEXISTANCE',
        'OPTIBAT_WATCHDOG': 'NO'
    },
    'ABG PALI': {
        'OPTIBAT_ON': 'YES',
        'Flag_Ready': 'NO',
        'Communication_ECS': 'NO',
        'Support_Flag_Copy': 'NO',
        'Macrostates_Flag_Copy': 'NO',
        'Resultexistance_Flag_Copy': 'NO',
        'OPTIBAT_WATCHDOG': 'NO'
    },
    'ANGUS': {
        'OPTIBAT_ON': 'NO',
        'Flag_Ready': 'NO',
        'Communication_ECS': 'NO',
        'Support_Flag_Copy': 'NO',
        'Macrostates_Flag_Copy': 'NO',
        'Resultexistance_Flag_Copy': 'NO',
        'OPTIBAT_WATCHDOG': 'NO',
        'Other_Flags': 'Purity_TA40_of_TAC_OPTIBAT, UV_Absorbance_OPTIBAT'
    },
    'CEMEX FM1 BALCONES': {
        'OPTIBAT_ON': 'YES',
        'Flag_Ready': 'YES',
        'Communication_ECS': 'YES',
        'Support_Flag_Copy': 'YES',
        'Macrostates_Flag_Copy': 'YES',
        'Resultexistance_Flag_Copy': 'YES',
        'OPTIBAT_WATCHDOG': 'YES'
    },
    'CRH LEMONA': {
        'OPTIBAT_ON': 'Kiln_OPTIBAT_ON',
        'Flag_Ready': 'KILN_OPTIBAT_READY / OPTIBAT_Ready_fromPLANT',
        'Communication_ECS': 'KILN_OPTIBAT_COMMUNICATION',
        'Support_Flag_Copy': 'NO',
        'Macrostates_Flag_Copy': 'NO',
        'Resultexistance_Flag_Copy': 'NO',
        'OPTIBAT_WATCHDOG': 'NO'
    },
    'MOLINS ALION COLOMBIA': {
        'OPTIBAT_ON': 'YES',
        'Flag_Ready': 'OPTIBAT_READY',
        'Communication_ECS': 'NO',
        'Support_Flag_Copy': 'YES',
        'Macrostates_Flag_Copy': 'YES',
        'Resultexistance_Flag_Copy': 'YES',
        'OPTIBAT_WATCHDOG': 'YES'
    },
    'MOLINS-BCN-BARCELONA': {
        'OPTIBAT_ON': 'YES',
        'Flag_Ready': 'NO',
        'Communication_ECS': 'KILN_OPTIBAT_COMMUNICATION',
        'Support_Flag_Copy': 'NO',
        'Macrostates_Flag_Copy': 'NO',
        'Resultexistance_Flag_Copy': 'NO',
        'OPTIBAT_WATCHDOG': 'NO'
    },
    'TITAN ALEXANDRIA CM7': {
        'OPTIBAT_ON': 'YES',
        'Flag_Ready': 'OPTIBAT_READY',
        'Communication_ECS': 'OPTIBAT_COMMUNICATION',
        'Support_Flag_Copy': 'NO',
        'Macrostates_Flag_Copy': 'NO',
        'Resultexistance_Flag_Copy': 'NO',
        'OPTIBAT_WATCHDOG': 'NO'
    },
    'TITAN ALEXANDRIA CM8': {
        'OPTIBAT_ON': 'YES',
        'Flag_Ready': 'OPTIBAT_READY',
        'Communication_ECS': 'OPTIBAT_COMMUNICATION',
        'Support_Flag_Copy': 'NO',
        'Macrostates_Flag_Copy': 'NO',
        'Resultexistance_Flag_Copy': 'NO',
        'OPTIBAT_WATCHDOG': 'NO'
    },
    'TITAN ALEXANDRIA CM9': {
        'OPTIBAT_ON': 'YES',
        'Flag_Ready': 'OPTIBAT_READY',
        'Communication_ECS': 'OPTIBAT_COMMUNICATION',
        'Support_Flag_Copy': 'YES',
        'Macrostates_Flag_Copy': 'YES',
        'Resultexistance_Flag_Copy': 'YES',
        'OPTIBAT_WATCHDOG': 'YES'
    },
    'TITAN-KOSJERIC-CM1': {
        'OPTIBAT_ON': 'YES',
        'Flag_Ready': 'OPTIBAT_READY',
        'Communication_ECS': 'OPTIBAT_COMMUNICATION',
        'Support_Flag_Copy': 'YES',
        'Macrostates_Flag_Copy': 'YES',
        'Resultexistance_Flag_Copy': 'YES',
        'OPTIBAT_WATCHDOG': 'NO'
    },
    'TITAN-KOSJERIC-KILN': {
        'OPTIBAT_ON': 'YES',
        'Flag_Ready': 'OPTIBAT_READY',
        'Communication_ECS': 'OPTIBAT_COMMUNICATION',
        'Support_Flag_Copy': 'YES',
        'Macrostates_Flag_Copy': 'YES',
        'Resultexistance_Flag_Copy': 'YES',
        'OPTIBAT_WATCHDOG': 'NO'
    },
    'TITAN-KOSJERIC-RM1': {
        'OPTIBAT_ON': 'NO',
        'Flag_Ready': 'NO',
        'Communication_ECS': 'NO',
        'Support_Flag_Copy': 'OPTIBAT_SUPPORT',
        'Macrostates_Flag_Copy': 'NO',
        'Resultexistance_Flag_Copy': 'NO',
        'OPTIBAT_WATCHDOG': 'NO'
    },
    'TITAN-PENNSUCO-FM3': {
        'OPTIBAT_ON': 'YES',
        'Flag_Ready': 'NO',
        'Communication_ECS': 'OPTIBAT_COMMUNICATION',
        'Support_Flag_Copy': 'NO',
        'Macrostates_Flag_Copy': 'NO',
        'Resultexistance_Flag_Copy': 'NO',
        'OPTIBAT_WATCHDOG': 'NO',
        'Other_Flags': '531OPTIBAT_BlaineInRange'
    },
    'TITAN-PENNSUCO-KILN': {
        'OPTIBAT_ON': '431OPTIBATON',
        'Flag_Ready': 'NO',
        'Communication_ECS': 'OPTIBAT_COMMUNICATION',
        'Support_Flag_Copy': 'NO',
        'Macrostates_Flag_Copy': 'NO',
        'Resultexistance_Flag_Copy': 'NO',
        'OPTIBAT_WATCHDOG': 'NO'
    },
    'TITAN-PENNSUCO-VRM': {
        'OPTIBAT_ON': 'OPTIBATON_OPC',
        'Flag_Ready': 'NO',
        'Communication_ECS': 'OPTIBAT_COMMUNICATION',
        'Support_Flag_Copy': 'NO',
        'Macrostates_Flag_Copy': 'NO',
        'Resultexistance_Flag_Copy': 'NO',
        'OPTIBAT_WATCHDOG': 'NO'
    },
    'TITAN-ROANOKE-KILN': {
        'OPTIBAT_ON': 'YES',
        'Flag_Ready': 'NO',
        'Communication_ECS': 'OPTIBAT_COMMUNICATION',
        'Support_Flag_Copy': 'NO',
        'Macrostates_Flag_Copy': 'NO',
        'Resultexistance_Flag_Copy': 'NO',
        'OPTIBAT_WATCHDOG': 'NO'
    },
    'TITAN-SHARR-CM2': {
        'OPTIBAT_ON': 'YES',
        'Flag_Ready': 'OPTIBAT_READY',
        'Communication_ECS': 'OPTIBAT_COMMUNICATION',
        'Support_Flag_Copy': 'YES',
        'Macrostates_Flag_Copy': 'YES',
        'Resultexistance_Flag_Copy': 'YES',
        'OPTIBAT_WATCHDOG': 'YES'
    },
    'TITAN-SHARR-KILN': {
        'OPTIBAT_ON': 'YES',
        'Flag_Ready': 'OPTIBAT_READY',
        'Communication_ECS': 'OPTIBAT_COMMUNICATION',
        'Support_Flag_Copy': 'YES',
        'Macrostates_Flag_Copy': 'YES',
        'Resultexistance_Flag_Copy': 'YES',
        'OPTIBAT_WATCHDOG': 'YES'
    },
    'TITAN-SHARR-RM2': {
        'OPTIBAT_ON': 'YES',
        'Flag_Ready': 'OPTIBAT_READY',
        'Communication_ECS': 'OPTIBAT_COMMUNICATION',
        'Support_Flag_Copy': 'YES',
        'Macrostates_Flag_Copy': 'YES',
        'Resultexistance_Flag_Copy': 'YES',
        'OPTIBAT_WATCHDOG': 'YES'
    },
    'TITAN-ROANOKE-FM10': {
        'OPTIBAT_ON': 'YES',
        'Flag_Ready': 'NO',
        'Communication_ECS': 'OPTIBAT_COMMUNICATION',
        'Support_Flag_Copy': 'NO',
        'Macrostates_Flag_Copy': 'NO',
        'Resultexistance_Flag_Copy': 'NO',
        'OPTIBAT_WATCHDOG': 'NO'
    },
    'TITAN-ROANOKE-RM1': {
        'OPTIBAT_ON': 'YES',
        'Flag_Ready': 'NO',
        'Communication_ECS': 'OPTIBAT_COMMUNICATION',
        'Support_Flag_Copy': 'NO',
        'Macrostates_Flag_Copy': 'NO',
        'Resultexistance_Flag_Copy': 'NO',
        'OPTIBAT_WATCHDOG': 'NO'
    }
}

# Create a DataFrame from the dictionary
df = pd.DataFrame.from_dict(client_flags, orient='index')
df.index.name = 'Cliente'

# Reset index to make Cliente a column
df = df.reset_index()

# Define the column order
columns_order = ['Cliente', 'OPTIBAT_ON', 'Flag_Ready', 'Communication_ECS', 
                 'Support_Flag_Copy', 'Macrostates_Flag_Copy', 
                 'Resultexistance_Flag_Copy', 'OPTIBAT_WATCHDOG']

# Add 'Other_Flags' column if it exists
if 'Other_Flags' in df.columns:
    columns_order.append('Other_Flags')

# Reorder columns
df = df[columns_order]

# Save to Excel
output_file = 'INFORME_FLAGS_CLIENTES_COMPLETO.xlsx'
df.to_excel(output_file, index=False, engine='openpyxl')

# Apply formatting
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# Load the workbook
wb = load_workbook(output_file)
ws = wb.active

# Define styles
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True)
center_alignment = Alignment(horizontal="center", vertical="center")
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Apply header formatting
for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center_alignment
    cell.border = thin_border

# Apply cell formatting and conditional formatting
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    for cell in row:
        cell.border = thin_border
        cell.alignment = center_alignment
        
        # Apply conditional formatting
        if cell.column > 1:  # Skip the first column (Cliente)
            if cell.value == 'YES':
                cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                cell.font = Font(color="006100")
            elif cell.value == 'NO':
                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                cell.font = Font(color="9C0006")
            elif cell.value and cell.value not in ['YES', 'NO']:
                cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                cell.font = Font(color="9C6500")

# Adjust column widths
for column in ws.columns:
    max_length = 0
    column_letter = column[0].column_letter
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = (max_length + 2) * 1.2
    ws.column_dimensions[column_letter].width = adjusted_width

# Save the formatted workbook
wb.save(output_file)

print(f"Excel file '{output_file}' has been created successfully!")
print(f"Total clients processed: {len(df)}")
print(f"Columns included: {', '.join(columns_order)}")

# Create a summary
summary = []
summary.append("SUMMARY OF FLAGS BY CLIENT:")
summary.append("=" * 50)
for client, flags in client_flags.items():
    summary.append(f"\n{client}:")
    for flag, value in flags.items():
        if value not in ['NO', None]:
            summary.append(f"  - {flag}: {value}")

# Save summary to text file
with open('FLAGS_SUMMARY.txt', 'w') as f:
    f.write('\n'.join(summary))

print(f"\nSummary saved to 'FLAGS_SUMMARY.txt'")