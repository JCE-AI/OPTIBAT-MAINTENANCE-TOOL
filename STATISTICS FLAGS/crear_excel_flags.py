import pandas as pd
import os
from datetime import datetime

# Datos de los clientes y sus flags
data = {
    'Cliente': [
        'ABG DALLA', 'ABG DHAR', 'ABG PALI', 'ANGUS', 'CEMEX FM1 BALCONES',
        'CRH LEMONA', 'MOLINS ALION COLOMBIA', 'MOLINS-BCN-BARACELONA',
        'TITAN ALEXANDRIA CM7', 'TITAN ALEXANDRIA CM8', 'TITAN ALEXANDRIA CM9',
        'TITAN-KOSJERIC-CM1', 'TITAN-KOSJERIC-KILN', 'TITAN-KOSJERIC-RM1',
        'TITAN-PENNSUCO-FM3', 'TITAN-PENNSUCO-KILN', 'TITAN-PENNSUCO-VRM',
        'TITAN-ROANOKE-FM10', 'TITAN-ROANOKE-KILN', 'TITAN-ROANOKE-RM1',
        'TITAN-SHARR-CM2', 'TITAN-SHARR-KILN', 'TITAN-SHARR-RM2'
    ],
    'OPTIBAT_ON': [
        'OPTIBAT_ON', 'OPTIBAT_ON', 'OPTIBAT_ON', '', 'OPTIBAT_ON',
        'OPTIBAT_ON', 'OPTIBAT_ON', 'OPTIBAT_ON',
        'OPTIBAT_ON', 'OPTIBAT_ON', 'OPTIBAT_ON',
        'OPTIBAT_ON', 'OPTIBAT_ON', 'OPTIBAT_ON',
        'OPTIBAT_ON', 'OPTIBAT_ON', 'OPTIBAT_ON',
        '', '', '',
        '', '', ''
    ],
    'Flag_Ready': [
        'OPTIBAT_READY', 'OPTIBAT_READY', 'OPTIBAT_READY', '', 'Flag_Ready',
        'OPTIBAT_Ready_fromPLANT', 'OPTIBAT_READY', 'OPTIBAT_READY',
        'OPTIBAT_READY', 'OPTIBAT_READY', 'OPTIBAT_READY',
        'OPTIBAT_READY', 'OPTIBAT_READY', 'OPTIBAT_READY',
        'OPTIBAT_READY', 'OPTIBAT_READY', 'OPTIBAT_READY',
        'OPTIBAT_READY', 'OPTIBAT_READY', 'OPTIBAT_READY',
        'OPTIBAT_READY', 'OPTIBAT_READY', 'OPTIBAT_READY'
    ],
    'Communication_ECS': [
        'OPTIBAT_COMMUNICATION', '', '', '', 'Communication_ECS',
        '', '', '',
        'OPTIBAT_COMMUNICATION', 'OPTIBAT_COMMUNICATION', 'OPTIBAT_COMMUNICATION',
        'OPTIBAT_COMMUNICATION', 'OPTIBAT_COMMUNICATION', 'OPTIBAT_COMMUNICATION',
        'KILN_OPTIBAT_COMMUNICATION', 'KILN_OPTIBAT_COMMUNICATION', 'OPTIBAT_COMMUNICATION',
        'Communication_Flag', 'Communication_Flag', 'Communication_Flag',
        '', '', ''
    ],
    'Support_Flag_Copy': [
        'OPTIBAT_SUPPORT', 'OPTIBAT_SUPPORT', 'OPTIBAT_SUPPORT', '', 'Support_Flag_Copy',
        'OPTIBAT_SUPPORT', 'OPTIBAT_SUPPORT', 'OPTIBAT_SUPPORT',
        'OPTIBAT_SUPPORT', 'OPTIBAT_SUPPORT', 'OPTIBAT_SUPPORT',
        'OPTIBAT_SUPPORT', 'OPTIBAT_SUPPORT', 'OPTIBAT_SUPPORT',
        'OPTIBAT_SUPPORT', 'OPTIBAT_SUPPORT', 'OPTIBAT_SUPPORT',
        '', '', '',
        '', '', ''
    ],
    'Macrostates_Flag_Copy': [
        'OPTIBAT_MACROSTATES', 'OPTIBAT_MACROSTATES', 'OPTIBAT_MACROSTATES', '', 'Macrostates_Flag_Copy',
        'OPTIBAT_MACROSTATES', 'OPTIBAT_MACROSTATES', 'OPTIBAT_MACROSTATES',
        'OPTIBAT_MACROSTATES', 'OPTIBAT_MACROSTATES', 'OPTIBAT_MACROSTATES',
        'OPTIBAT_MACROSTATES', 'OPTIBAT_MACROSTATES', 'OPTIBAT_MACROSTATES',
        'OPTIBAT_MACROSTATES', 'OPTIBAT_MACROSTATES', 'OPTIBAT_MACROSTATES',
        '', '', '',
        '', '', ''
    ],
    'Resultexistance_Flag_Copy': [
        'OPTIBAT_RESULTEXISTANCE', 'OPTIBAT_RESULTEXISTANCE', 'OPTIBAT_RESULTEXISTANCE', '', 'Resultexistance_Flag_Copy',
        'OPTIBAT_RESULTEXISTANCE', 'OPTIBAT_RESULTEXISTANCE', 'OPTIBAT_RESULTEXISTANCE',
        '', '', '',
        'OPTIBAT_RESULTEXISTANCE', 'OPTIBAT_RESULTEXISTANCE', 'OPTIBAT_RESULTEXISTANCE',
        'OPTIBAT_RESULTEXISTANCE', 'OPTIBAT_RESULTEXISTANCE', 'OPTIBAT_RESULTEXISTANCE',
        '', '', '',
        '', '', ''
    ],
    'OPTIBAT_WATCHDOG': [
        'OPTIBAT_WATCHDOG', '', '', '', 'OPTIBAT_WATCHDOG',
        '', '', '',
        '', '', '',
        '', '', '',
        '', '', '',
        '', '', '',
        '', '', ''
    ]
}

# Crear DataFrame
df = pd.DataFrame(data)

# Calcular total de flags por cliente
df['Total_Flags'] = df.iloc[:, 1:].apply(lambda row: sum(1 for x in row if x != ''), axis=1)

# Crear archivo Excel con formato
output_path = r'C:\Users\JuanCruz\Desktop_Local\mtto streamlit\STATISTICS FLAGS\INFORME_FLAGS_CLIENTES.xlsx'

# Crear Excel writer con openpyxl
with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
    df.to_excel(writer, sheet_name='Flags por Cliente', index=False)
    
    # Obtener el worksheet
    worksheet = writer.sheets['Flags por Cliente']
    
    # Ajustar ancho de columnas
    worksheet.column_dimensions['A'].width = 25  # Cliente
    for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
        worksheet.column_dimensions[col].width = 30
    worksheet.column_dimensions['I'].width = 12  # Total
    
    # Aplicar formato a encabezados
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    center_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                        top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Formato para encabezados
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = thin_border
    
    # Formato para datos
    green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Colorear celdas según contenido
            if cell.column > 1 and cell.column < 9:  # Columnas de flags
                if cell.value and cell.value != '':
                    cell.fill = green_fill
                else:
                    cell.fill = red_fill
            
            # Formato especial para columna Total
            if cell.column == 9:
                if cell.value == 7:
                    cell.fill = PatternFill(start_color='00B050', end_color='00B050', fill_type='solid')
                    cell.font = Font(bold=True, color='FFFFFF')
                elif cell.value >= 5:
                    cell.fill = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')
                elif cell.value >= 3:
                    cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                else:
                    cell.fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
                    cell.font = Font(color='FFFFFF')

# Crear hoja de resumen
with pd.ExcelWriter(output_path, engine='openpyxl', mode='a') as writer:
    # Resumen estadístico
    resumen_data = {
        'Métrica': [
            'Total de Clientes',
            'Clientes con todas las flags (7/7)',
            'Clientes sin flags OPTIBAT',
            'Promedio de flags por cliente',
            'Flag más común',
            'Flag menos común'
        ],
        'Valor': [
            23,
            2,
            1,
            f"{df['Total_Flags'].mean():.1f}",
            'Flag_Ready (23 clientes)',
            'OPTIBAT_WATCHDOG (2 clientes)'
        ]
    }
    
    df_resumen = pd.DataFrame(resumen_data)
    df_resumen.to_excel(writer, sheet_name='Resumen', index=False)
    
    # Formato para hoja de resumen
    worksheet = writer.sheets['Resumen']
    worksheet.column_dimensions['A'].width = 35
    worksheet.column_dimensions['B'].width = 25
    
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = thin_border
    
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
        for cell in row:
            cell.border = thin_border
            cell.alignment = center_alignment

print(f"Archivo Excel creado exitosamente en: {output_path}")
print(f"Total de clientes analizados: 23")
print(f"Clientes con implementación completa (7/7 flags): ABG DALLA, CEMEX FM1 BALCONES")
print(f"Cliente sin flags OPTIBAT: ANGUS")